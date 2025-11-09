# src/panviral_hdf/rf/pipeline.py
# -*- coding: utf-8 -*-
"""
End-to-end RF pipeline for HDF vs non-HDF

"""

from __future__ import annotations

import os
import re
import gc
import sys
import gzip
import json
import time
import math
import glob
import uuid
import joblib
import random
import string
import warnings
import platform
import traceback
import datetime as dt
from collections import Counter, defaultdict
from io import StringIO
from typing import List, Dict, Any

import numpy as np
import pandas as pd

# ML & metrics
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.model_selection import StratifiedShuffleSplit, RepeatedStratifiedKFold, GridSearchCV
from sklearn.feature_selection import RFE
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import (
    accuracy_score, balanced_accuracy_score, roc_auc_score, precision_recall_curve,
    average_precision_score, confusion_matrix, cohen_kappa_score, matthews_corrcoef,
    precision_score, recall_score, f1_score, brier_score_loss
)

# Optional: imbalanced-learn
try:
    from imblearn.pipeline import Pipeline as ImbPipeline
    from imblearn.over_sampling import SMOTE
    _HAS_IMB = True
except Exception:
    _HAS_IMB = False

# Optional: read R objects
try:
    import pyreadr
    _HAS_PYREADR = True
except Exception:
    _HAS_PYREADR = False

# Optional: mapping SYMBOL/ALIAS → ENSEMBL
try:
    import mygene
    _HAS_MYG = True
except Exception:
    _HAS_MYG = False

# Excel + plotting
import openpyxl
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from upsetplot import from_contents, UpSet
    _HAS_UPSET = True
except Exception:
    _HAS_UPSET = False

warnings.filterwarnings("ignore")

# -----------------------------
# User paths & constants  
# -----------------------------
ROOT_RAW    = "/media/mohadeseh/d2987156-83a1-4537-b507-30f08b63b454/Naseri/FinalFolder/HRF/Emtahan_dobare_HDF/"
ROOT_PUB    = "/media/mohadeseh/d2987156-83a1-4537-b507-30f08b63b454/Naseri/FinalFolder/HRF/Publications"
FEATURES_CSV= "/media/mohadeseh/d2987156-83a1-4537-b507-30f08b63b454/Naseri/Turtle/Desktop/Masterarbeit/Features/combinedFinal.csv"
VIRUSES     = ["Denv","IAV","Sars_Cov_2","Zika"]

TOP_N       = 1000
TARGET_N    = 1000
OUT_SUBDIR  = "MAIC_top_with_ensembl"
BAL_SUBDIR  = "balanced_from_bottom"

POS_CLASS   = "HDF"     # کلاس مثبت
NEG_CLASS   = "HRF"     # non-HDF (Low-MAIC)
random_seed = 113
np.random.seed(random_seed)
random.seed(random_seed)


def _message(*args):
    print(*args, flush=True)

def try_step(func, tag):
    try:
        return func()
    except Exception as e:
        raise RuntimeError(f"[{tag}] {e}")

def sanitize_sheet(x: str) -> str:
    x = re.sub(r"[\[\]\*\?\/\\:]", "_", x)
    x = x.strip()
    x = x[:31]
    return x or "Sheet"

def make_unique_sheet(wb, name: str) -> str:
    n = sanitize_sheet(name)
    existing = [ws.title for ws in wb.worksheets]
    if n not in existing:
        return n
    i = 2
    base = n[:28]
    while True:
        cand = sanitize_sheet(f"{base}_{i}")
        if cand not in existing:
            return cand
        i += 1


def map_symbols_to_ensembl_first(symbols):
    if not _HAS_MYG:
        raise RuntimeError("mygene نصب نشده است؛ برای نگاشت SYMBOL→ENSEMBL لطفاً mygene را نصب کنید.")
    mg = mygene.MyGeneInfo()
    syms = [str(s).strip() if s is not None else "" for s in symbols]
    uniq_syms = list(dict.fromkeys(syms))
    res1 = mg.querymany(uniq_syms, scopes="symbol", fields="ensembl.gene", species="human",
                        as_dataframe=False, returnall=False, verbose=False, skip_missing=True, size=50)
    map1 = defaultdict(list)
    for r in res1:
        if 'notfound' in r and r['notfound']:
            continue
        ensg = r.get('ensembl', {})
        if isinstance(ensg, dict) and 'gene' in ensg:
            vals = [ensg['gene']]
        elif isinstance(ensg, list):
            vals = []
            for it in ensg:
                v = it.get('gene')
                if v: vals.append(v)
        else:
            vals = []
        vals = [v for v in vals if isinstance(v, str) and v.startswith("ENSG")]
        if vals:
            map1[r['query']].extend(sorted(set(vals)))

    missing = [s for s in uniq_syms if len(map1[s]) == 0]
    if missing:
        res2 = mg.querymany(missing, scopes="alias", fields="ensembl.gene", species="human",
                            as_dataframe=False, returnall=False, verbose=False, skip_missing=True, size=50)
        for r in res2:
            if 'notfound' in r and r['notfound']:
                continue
            ensg = r.get('ensembl', {})
            vals = []
            if isinstance(ensg, dict) and 'gene' in ensg:
                vals = [ensg['gene']]
            elif isinstance(ensg, list):
                for it in ensg:
                    v = it.get('gene')
                    if v: vals.append(v)
            vals = [v for v in vals if isinstance(v, str) and v.startswith("ENSG")]
            if vals:
                map1[r['query']].extend(sorted(set(vals)))

    out = []
    for s in syms:
        arr = list(dict.fromkeys(map1.get(s, [])))
        out.append(arr[0] if arr else np.nan)
    return np.array(out, dtype=object)


def read_features_from_path(path):
    ext = os.path.splitext(path)[1].lower().strip(".")
    if ext == "csv":
        dt = pd.read_csv(path)
    elif ext in ["tsv","txt"]:
        dt = pd.read_csv(path, sep="\t")
    elif ext in ["rds","rdata","rda"]:
        if not _HAS_PYREADR:
            raise RuntimeError("pyreadr نصب نشده است؛ برای خواندن RDS/RData لازم است.")
        res = pyreadr.read_r(path)
        objnames = list(res.keys())
        tgt = "Data_2" if "Data_2" in objnames else objnames[0]
        dt = res[tgt]
        if not isinstance(dt, pd.DataFrame):
            dt = pd.DataFrame(dt)
    else:
        raise RuntimeError(f"Unsupported features format: {ext}")
    cand = [c for c in ["Ensembl_ID","ensembl_id","ENSEMBL","Ensembl","ensembl"] if c in dt.columns]
    if cand:
        dt = dt.rename(columns={cand[0]:"Ensembl_ID"})
    if "Ensembl_ID" not in dt.columns:
        raise RuntimeError("ستون 'Ensembl_ID' در فیچرها نیست.")
    return dt

# --- nearZeroVar شبیه caret ---
def near_zero_var(df: pd.DataFrame):
    nzv = []
    metrics = []
    n = len(df)
    for c in df.columns:
        s = df[c]
        vals = s.astype("object")
        unique_vals, counts = np.unique(vals[~pd.isna(vals)], return_counts=True)
        pct_unique = (len(unique_vals) / max(1, n)) * 100.0
        if len(counts) == 0:
            nzv_flag = True
            freq_ratio = np.inf
        elif len(counts) == 1:
            nzv_flag = True
            freq_ratio = np.inf
        else:
            counts_sorted = np.sort(counts)[::-1]
            freq_ratio = counts_sorted[0] / max(1, counts_sorted[1])
            nzv_flag = (pct_unique <= 10.0 and freq_ratio > 19.0)
        nzv.append(nzv_flag)
        metrics.append((pct_unique, freq_ratio))
    out = pd.DataFrame(metrics, columns=["percentUnique","freqRatio"], index=df.columns)
    out["nzv"] = nzv
    return out


def find_correlation(corr, threshold=0.70):
    if isinstance(corr, pd.DataFrame):
        C = corr.values.copy()
        cols = list(corr.columns)
    else:
        C = corr.copy()
        cols = [str(i) for i in range(C.shape[0])]
    np.fill_diagonal(C, 0.0)
    to_drop = set()
    while True:
        if C.size == 0:
            break
        max_corr = np.max(np.abs(C))
        if max_corr < threshold:
            break
        idx = np.unravel_index(np.argmax(np.abs(C)), C.shape)
        i, j = idx
        mean_i = np.mean(np.abs(C[i, :]))
        mean_j = np.mean(np.abs(C[j, :]))
        drop_idx = i if mean_i >= mean_j else j
        to_drop.add(cols[drop_idx])
        keep = [k for k in range(C.shape[0]) if k != drop_idx]
        C = C[np.ix_(keep, keep)]
        cols = [cols[k] for k in keep]
    return list(to_drop)


def compute_metrics_safe(y_true, y_pred, y_prob, pos_label="HDF"):
    labels = [pos_label, NEG_CLASS] if pos_label in ["HDF","HRF"] else sorted(list(set(y_true)))
    cm = confusion_matrix(y_true, y_pred, labels=labels)
    pos_idx = 0
    neg_idx = 1 if cm.shape == (2,2) else 0
    TP = cm[pos_idx, pos_idx] if cm.size else 0
    TN = cm[neg_idx, neg_idx] if cm.size else 0
    FP = cm[neg_idx, pos_idx] if cm.size else 0
    FN = cm[pos_idx, neg_idx] if cm.size else 0

    acc = accuracy_score(y_true, y_pred) if len(y_true) else np.nan
    bal_acc = balanced_accuracy_score(y_true, y_pred) if len(y_true) else np.nan
    try:
        auc = roc_auc_score((np.array(y_true)==pos_label).astype(int), y_prob) if len(np.unique(y_true))>1 else np.nan
    except Exception:
        auc = np.nan

    prec = precision_score(y_true, y_pred, pos_label=pos_label, zero_division=0)
    reca = recall_score(y_true, y_pred, pos_label=pos_label, zero_division=0)
    f1   = f1_score(y_true, y_pred, pos_label=pos_label, zero_division=0)
    kap  = cohen_kappa_score(y_true, y_pred) if len(y_true) else np.nan
    mcc  = matthews_corrcoef(y_true, y_pred) if len(y_true) else np.nan
    brier= brier_score_loss((np.array(y_true)==pos_label).astype(int), y_prob) if len(y_true) else np.nan

    # Bootstrap CI برای AUC
    auc_lower = auc_upper = np.nan
    if not np.isnan(auc):
        rng = np.random.RandomState(2025)
        B = 1000
        idx = np.arange(len(y_true))
        boot = []
        for _ in range(B):
            bs = rng.choice(idx, size=len(idx), replace=True)
            yb = np.array(y_true)[bs]
            pb = np.array(y_prob)[bs]
            if len(np.unique(yb))<2:
                continue
            try:
                boot.append(roc_auc_score((yb==pos_label).astype(int), pb))
            except Exception:
                pass
        if boot:
            auc_lower, auc_upper = np.percentile(boot, [2.5,97.5])
    return dict(tp=int(TP), tn=int(TN), fp=int(FP), fn=int(FN), acc=acc, bal_acc=bal_acc,
                auc=auc, auc_lower=auc_lower, auc_upper=auc_upper,
                prec=prec, reca=reca, f1=f1, kap=kap, mcc=mcc, brier=brier)

def read_table_smart(path):
    return pd.read_csv(path, sep="\t", header=0, dtype=str, quoting=3, engine="python")

def save_tsv(df, out_path, index=False):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    df.to_csv(out_path, sep="\t", index=index)

def save_csv(df, out_path, index=False):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    df.to_csv(out_path, index=index)

def pretty_name(base_stub: str) -> str:
    s = re.sub(r"^filtered_HDF_", "", base_stub, flags=re.I)
    s = re.sub(r"_bottom\d+_HRF__top\d+_HDF$", "", s, flags=re.I)
    s = re.sub(r"_balanced_top_vs_bottom$", "", s, flags=re.I)
    return s

# ============================================================
def process_one_hdf_file(file_path, out_dir, top_n=1000):
    _message(">>> خواندن فایل:", file_path)
    dt = read_table_smart(file_path)
    if dt.shape[0] == 0:
        _message("⚠️ فایل خالی:", file_path); return
    nms = [c.lower() for c in dt.columns]
    gene_col_idx  = [i for i,n in enumerate(nms) if n=="gene"]
    if not gene_col_idx:
        raise RuntimeError("ستون 'gene' پیدا نشد: " + file_path)
    gene_col = dt.columns[gene_col_idx[0]]
    score_col_idx = [i for i,n in enumerate(nms) if n=="maic_score"]
    if score_col_idx:
        sc = dt.columns[score_col_idx[0]]
        dt[sc] = pd.to_numeric(dt[sc], errors="coerce")
        dt = dt.sort_values(sc, ascending=False, na_position="last").reset_index(drop=True)
    dt["rank_global"] = np.arange(1, dt.shape[0]+1)

    def map_and_write(sub_dt, label_suffix):
        syms = sub_dt[gene_col].astype(str).str.strip().tolist()
        ens_first = map_symbols_to_ensembl_first(syms)
        out = sub_dt.copy()
        out.insert(1, "ensembl_id", ens_first)
        removed_na = int(pd.isna(out["ensembl_id"]).sum())
        out = out[~pd.isna(out["ensembl_id"])].copy()
        os.makedirs(out_dir, exist_ok=True)
        base = os.path.basename(file_path)
        base_noext = re.sub(r"\.txt$", "", base, flags=re.I)
        out_file = os.path.join(out_dir, f"{base_noext}_{label_suffix}{top_n}_with_ensembl.tsv")
        save_tsv(out, out_file)
        _message(f"    ↳ {label_suffix.upper()}: {out.shape[0]} ردیف ذخیره شد (حذف {removed_na} بدون Ensembl): {out_file}")

    N_top = min(top_n, dt.shape[0]);   top_dt = dt.iloc[:N_top, :].copy();      top_dt["rank_top"] = np.arange(1, top_dt.shape[0]+1); map_and_write(top_dt, "top")
    N_bot = min(top_n, dt.shape[0]);   bottom_dt = dt.tail(N_bot).copy();       bottom_dt["rank_bottom"] = np.arange(1, bottom_dt.shape[0]+1); map_and_write(bottom_dt, "bottom")

def step_build_top_bottom(viruses=VIRUSES, root_raw=ROOT_RAW, top_n=TOP_N, out_subdir=OUT_SUBDIR):
    for virus in viruses:
        virus_dir = os.path.join(root_raw, virus)
        if not os.path.isdir(virus_dir):
            _message("⏭️ پوشه پیدا نشد:", virus_dir); continue
        _message("====== ویروس:", virus, "======")
        hdf_files = [p for p in glob.glob(os.path.join(virus_dir, "filtered_H*.txt")) if os.path.isfile(p)]
        if not hdf_files:
            _message("⏭️ filtered_H*.txt یافت نشد در:", virus_dir); continue
        out_dir = os.path.join(virus_dir, out_subdir)
        for fp in hdf_files:
            process_one_hdf_file(fp, out_dir=out_dir, top_n=top_n)
    _message("✅ ساخت فایل‌های Top/Bottom تمام شد.")

# ============================================================
def pick_hdf_top_files(virus, root=ROOT_RAW, subdir=OUT_SUBDIR, prefer_all_first=True):
    dir_top = os.path.join(root, virus, subdir)
    if not os.path.isdir(dir_top):
        raise RuntimeError("پوشه وجود ندارد: " + dir_top)
    files = [p for p in glob.glob(os.path.join(dir_top, "filtered_HDF_*_top*_with_ensembl.tsv")) if os.path.isfile(p)]
    if not files:
        raise RuntimeError(f"هیچ فایل *_top*_with_ensembl.tsv در {dir_top} پیدا نشد.")
    if prefer_all_first:
        files_all = [f for f in files if re.search(r"All_Categories", os.path.basename(f), flags=re.I)]
        files = files_all + [f for f in files if f not in files_all]
    seen = set(); uniq = []
    for f in files:
        if f not in seen:
            uniq.append(f); seen.add(f)
    return uniq

def label_top_bottom_exact_from_raw(hdf_top_file, features_path, n_each=1000):
    dir_top = os.path.dirname(hdf_top_file)
    dir_raw = os.path.dirname(dir_top)
    base = re.sub(r"_top\d+_with_ensembl\.tsv$", "", os.path.basename(hdf_top_file), flags=re.I)
    raw_guess = os.path.join(dir_raw, base + ".txt")
    if not os.path.isfile(raw_guess):
        pats = re.compile(rf"^{re.escape(base)}\.txt$", flags=re.I)
        cands = [p for p in glob.glob(os.path.join(dir_raw, "*.txt")) if pats.search(os.path.basename(p))]
        if not cands:
            raise RuntimeError("فایل خام پیدا نشد: " + raw_guess)
        raw_guess = cands[0]
    dt = read_table_smart(raw_guess)
    nms = [c.lower() for c in dt.columns]
    gene_col_idx = [i for i,n in enumerate(nms) if n=="gene"]
    if not gene_col_idx:
        raise RuntimeError("ستون 'gene' در فایل خام نیست: " + raw_guess)
    gene_col = dt.columns[gene_col_idx[0]]
    score_col_idx = [i for i,n in enumerate(nms) if n=="maic_score"]
    if score_col_idx:
        sc = dt.columns[score_col_idx[0]]
        dt[sc] = pd.to_numeric(dt[sc], errors="coerce")
        dt = dt.sort_values(sc, ascending=False, na_position="last").reset_index(drop=True)

    syms = dt[gene_col].astype(str).str.strip().tolist()
    ens_first = map_symbols_to_ensembl_first(syms)
    dt["ensembl_id"] = ens_first
    dt = dt[~pd.isna(dt["ensembl_id"])].copy()
    ordered_ens = pd.unique(dt["ensembl_id"]).tolist()

    feat = read_features_from_path(features_path).copy()
    present = pd.unique(feat["Ensembl_ID"].astype(str).str.strip()).tolist()

    def pick_n_in_order(vec, present_ids, exclude_ids=None, n_target=1000, from_tail=False):
        exclude_ids = set(exclude_ids or [])
        vec_iter = reversed(vec) if from_tail else vec
        chosen = []
        for gid in vec_iter:
            if gid in present_ids and gid not in exclude_ids and gid not in chosen:
                chosen.append(gid)
                if len(chosen) == n_target:
                    break
        return list(reversed(chosen)) if from_tail else chosen

    top_ids    = pick_n_in_order(ordered_ens, present, n_target=n_each, from_tail=False)            # HDF
    bottom_ids = pick_n_in_order(ordered_ens, present, exclude_ids=top_ids, n_target=n_each, from_tail=True)  # HRF

    if len(top_ids) < n_each:
        _message(f"فقط {len(top_ids)} آیتم از {n_each} تای اول در فیچرها پیدا شد.")
    if len(bottom_ids) < n_each:
        _message(f"فقط {len(bottom_ids)} آیتم از {n_each} تای آخر در فیچرها پیدا شد.")

    feat = feat.copy()
    feat["Class"] = np.nan
    if top_ids:
        feat.loc[feat["Ensembl_ID"].isin(top_ids), "Class"] = "HDF"
    if bottom_ids:
        feat.loc[feat["Ensembl_ID"].isin(bottom_ids), "Class"] = "HRF"
    feat_labeled = feat[~feat["Class"].isna()].copy()

    counts = feat_labeled["Class"].value_counts().reset_index()
    counts.columns = ["Class","N"]
    counts = counts.sort_values("N", ascending=False).reset_index(drop=True)

    return dict(data=feat_labeled, counts=counts, picked={"HDF_top":top_ids,"HRF_bottom":bottom_ids}, raw_file=raw_guess)

def step_label_and_balance(viruses=VIRUSES, features_csv=FEATURES_CSV, out_subdir=OUT_SUBDIR, bal_subdir=BAL_SUBDIR, target_n=TARGET_N):
    for virus in viruses:
        _message("====== Labeling →", virus, "======")
        try:
            top_files = pick_hdf_top_files(virus, root=ROOT_RAW, subdir=out_subdir)
        except Exception as e:
            _message("⏭️", e); continue
        for hdf_file in top_files:
            _message("→ پردازش فایل:", os.path.basename(hdf_file))
            res = label_top_bottom_exact_from_raw(hdf_top_file=hdf_file, features_path=features_csv, n_each=target_n)
            out_dir = os.path.join(os.path.dirname(hdf_file), bal_subdir)
            os.makedirs(out_dir, exist_ok=True)
            base_stub = re.sub(r"_top\d+_with_ensembl\.tsv$", "", os.path.basename(hdf_file), flags=re.I)
            out_csv = os.path.join(out_dir, f"{base_stub}_bottom{target_n}_HRF__top{target_n}_HDF.csv")
            res["data"].to_csv(out_csv, index=False)
            _message(res["counts"].to_string(index=False))
    _message("============================================================")
    _message("Start ML")


def run_ml_and_export_csv(df, balanced_csv_path, pos_class="HDF"):
    df = df.copy()
    if "Class" not in df.columns:
        raise RuntimeError("ستون Class وجود ندارد.")
    id_candidates = ["Ensembl_ID","ENSEMBL_ID","EnsemblId","ensembl_id"]
    id_col = next((c for c in id_candidates if c in df.columns), None)
    if id_col is None:
        raise RuntimeError("ستون شناسه‌ی Ensembl_ID در دیتافریم یافت نشد.")
    df["Class"] = df["Class"].astype(str)
    all_levels = ["HDF","HRF"]
    ord_levels = [pos_class] + [x for x in all_levels if x != pos_class]
    df["Class"] = pd.Categorical(df["Class"], categories=ord_levels, ordered=True)

    target_col = "Class"
    pred_cols  = [c for c in df.columns if c not in [id_col, target_col]]

    for c in pred_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    sss = StratifiedShuffleSplit(n_splits=1, test_size=0.20, random_state=random_seed)
    idx_train, idx_test = next(sss.split(df[pred_cols], df[target_col]))
    train_df = df.iloc[idx_train].reset_index(drop=True)
    test_df  = df.iloc[idx_test].reset_index(drop=True)

    nzv_metrics = near_zero_var(train_df[pred_cols].copy())
    if "nzv" not in nzv_metrics.columns:
        raise RuntimeError("[NZV] ستون 'nzv' در خروجی nearZeroVar یافت نشد.")
    keep_pred_nzv = [c for c in pred_cols if not bool(nzv_metrics.loc[c,"nzv"])]
    if not keep_pred_nzv:
        raise RuntimeError("[NZV] همهٔ پیش‌بین‌ها NZV شدند!")

    base_cols = [id_col, target_col]
    keep_cols1 = base_cols + keep_pred_nzv
    train_df = train_df[keep_cols1].copy()
    test_df  = test_df[keep_cols1].copy()
    pred_cols = [c for c in keep_cols1 if c not in base_cols]
    if not pred_cols:
        raise RuntimeError("[PRED_COLS] پس از NZV هیچ ویژگی‌ای باقی نماند.")

    preproc = Pipeline(steps=[
        ("imputer", SimpleImputer(strategy="median")),
        ("scaler", StandardScaler(with_mean=True, with_std=True))
    ])

    X_train = preproc.fit_transform(train_df[pred_cols])
    X_test  = preproc.transform(test_df[pred_cols])

    y_train = train_df[target_col].astype(str).values
    y_test  = test_df[target_col].astype(str).values

    classes = ["HDF","HRF"]
    cnt_train = dict(Counter(pd.Categorical(y_train, categories=classes)))
    cnt_test  = dict(Counter(pd.Categorical(y_test,  categories=classes)))
    _message(f"Label counts → Train: HDF={cnt_train.get('HDF',0)}, HRF={cnt_train.get('HRF',0)} | Test: HDF={cnt_test.get('HDF',0)}, HRF={cnt_test.get('HRF',0)}")

    p = X_train.shape[1]
    sizes_grid = [s for s in [32,64,128,256,512] if s <= p]
    if not sizes_grid:
        sizes_grid = list(range(1, min(10, max(1, p))+1))

    def rfe_select(X, y, sizes):
        best_score = -np.inf
        best_vars = None
        feat_names = np.array(pred_cols)
        for s in sizes:
            est = RandomForestClassifier(
                n_estimators=500,
                random_state=random_seed,
                n_jobs=-1,
                class_weight=None
            )
            rfe = RFE(estimator=est, n_features_to_select=s, step=1)
            cv = RepeatedStratifiedKFold(n_splits=5, n_repeats=2, random_state=random_seed)
            scores = []
            for tr, va in cv.split(X, y):
                X_tr, X_va = X[tr], X[va]
                y_tr, y_va = y[tr], y[va]
                try:
                    rfe.fit(X_tr, y_tr)
                    X_tr_sel = rfe.transform(X_tr)
                    X_va_sel = rfe.transform(X_va)
                    clf = RandomForestClassifier(
                        n_estimators=500,
                        random_state=random_seed,
                        n_jobs=-1
                    )
                    clf.fit(X_tr_sel, y_tr)
                    prob = clf.predict_proba(X_va_sel)[:, 0 if POS_CLASS==clf.classes_[0] else list(clf.classes_).index(POS_CLASS)]
                    sc = roc_auc_score((y_va==POS_CLASS).astype(int), prob)
                    scores.append(sc)
                except Exception:
                    pass
            if scores:
                sc_mean = float(np.mean(scores))
                if sc_mean > best_score:
                    best_score = sc_mean
                    rfe.fit(X, y)
                    support = rfe.get_support()
                    best_vars = feat_names[support].tolist()
        if best_vars is None or len(best_vars)==0:
            best_vars = feat_names.tolist()
        return best_vars, len(best_vars)

    opt_vars, opt_size = rfe_select(X_train, y_train, sizes_grid)

    sel_vars = opt_vars.copy()
    if len(sel_vars) > 1:
        idx_sel = [pred_cols.index(v) for v in sel_vars]
        cor_sel = pd.DataFrame(np.corrcoef(X_train[:, idx_sel].T), index=sel_vars, columns=sel_vars)
        drop_sel = find_correlation(cor_sel, threshold=0.70)
        if drop_sel:
            sel_vars = [v for v in sel_vars if v not in drop_sel]
        if not sel_vars:
            sel_vars = opt_vars

    tab_train = Counter(y_train)
    imb_ratio = max(tab_train.values()) / max(1, min(tab_train.values()))
    use_smote = (imb_ratio >= 1.5) and _HAS_IMB

    p_sel = len(sel_vars)
    mtry_grid = sorted(set([max(1, int(round(math.sqrt(p_sel) * f))) for f in [0.5,1,2,3]]))
    max_features_grid = [min(p_sel, v) for v in mtry_grid]

    if use_smote:
        pipe = ImbPipeline(steps=[
            ("smote", SMOTE(random_state=random_seed, n_jobs=-1)),
            ("rf", RandomForestClassifier(n_estimators=500, random_state=random_seed, n_jobs=-1))
        ])
        param_grid = {"rf__max_features": max_features_grid}
    else:
        pipe = Pipeline(steps=[
            ("rf", RandomForestClassifier(n_estimators=500, random_state=random_seed, n_jobs=-1))
        ])
        param_grid = {"rf__max_features": max_features_grid}

    rskf = RepeatedStratifiedKFold(n_splits=10, n_repeats=3, random_state=random_seed)
    sel_idx_train = [pred_cols.index(v) for v in sel_vars]
    X_train_sel = X_train[:, sel_idx_train]
    X_test_sel  = X_test[:,  sel_idx_train]

    gs = GridSearchCV(
        estimator=pipe,
        param_grid=param_grid,
        scoring="roc_auc",
        cv=rskf,
        n_jobs=-1,
        verbose=0,
        refit=True
    )
    gs.fit(X_train_sel, y_train)

    best_model = gs.best_estimator_

    pred_class = best_model.predict(X_test_sel)
    classes_model = getattr(best_model[-1], "classes_", None) if use_smote else getattr(best_model.named_steps["rf"], "classes_", None)
    if classes_model is None:
        classes_model = np.unique(y_train)
    pos_index = list(classes_model).index(pos_class) if pos_class in classes_model else 0
    pred_prob = best_model.predict_proba(X_test_sel)[:, pos_index]

    M = compute_metrics_safe(y_test, pred_class, pred_prob, pos_label=pos_class)

    base_name = os.path.basename(balanced_csv_path)
    base_stub = re.sub(r"(_balanced_top_vs_bottom|_bottom\d+_HRF__top\d+_HDF)\.csv$", "", base_name, flags=re.I)
    out_dir   = os.path.join(os.path.dirname(balanced_csv_path), "ML_results", base_stub)
    os.makedirs(out_dir, exist_ok=True)

    class_counts_df = pd.DataFrame({
        "Split": ["Train","Train","Test","Test"],
        "Class": ["HDF","HRF","HDF","HRF"],
        "N":     [cnt_train.get("HDF",0), cnt_train.get("HRF",0), cnt_test.get("HDF",0), cnt_test.get("HRF",0)]
    })
    class_counts_df["Percent"] = [round(100*c/sum([cnt_train.get("HDF",0),cnt_train.get("HRF",0)]),2) if i<2 else
                                  round(100*c/sum([cnt_test.get("HDF",0),cnt_test.get("HRF",0)]),2)
                                  for i,c in enumerate(class_counts_df["N"].tolist())]
    save_csv(class_counts_df, os.path.join(out_dir, "class_counts_train_test.csv"))

    n_test = X_test_sel.shape[0]
    pos_in_test = int((np.array(y_test)==pos_class).sum())
    neg_in_test = n_test - pos_in_test
    metrics_df = pd.DataFrame({
        "Metric": ["N_test","Pos_in_test","Neg_in_test",
                   "Accuracy","Balanced_Accuracy","AUC","AUC_95CI_Lower","AUC_95CI_Upper",
                   "AUPRC","Sensitivity","Specificity","Precision","Recall","F1","Kappa","MCC","Brier_Score"],
        "Value":  [n_test, pos_in_test, neg_in_test,
                   M["acc"], M["bal_acc"], M["auc"], M["auc_lower"], M["auc_upper"],
                   average_precision_score((np.array(y_test)==pos_class).astype(int), pred_prob),
                   M["reca"],
                   M["tn"]/(M["tn"]+M["fp"]) if (M["tn"]+M["fp"])>0 else np.nan,
                   M["prec"], M["reca"], M["f1"], M["kap"], M["mcc"], M["brier"]]
    })
    save_csv(metrics_df, os.path.join(out_dir, "metrics_test.csv"))

    lvl = [pos_class, NEG_CLASS]
    cm = confusion_matrix(y_test, pred_class, labels=lvl)
    cm_full = pd.DataFrame({
        "Reference": np.repeat(lvl, len(lvl)),
        "Prediction": lvl * len(lvl),
        "Count": cm.flatten().astype(int)
    })
    save_csv(cm_full, os.path.join(out_dir, "confusion_matrix_test.csv"))

    rf_final = best_model[-1] if use_smote else best_model.named_steps["rf"]
    importances = getattr(rf_final, "feature_importances_", None)
    if importances is None:
        imp_df = pd.DataFrame({"Feature": sel_vars, "Importance": np.nan})
    else:
        imp_df = pd.DataFrame({"Feature": sel_vars, "Importance": importances})
        imp_df = imp_df.sort_values("Importance", ascending=False)
    top_imp = imp_df.dropna().head(30).copy()
    save_csv(top_imp, os.path.join(out_dir, "top_features.csv"))

    preds_df = pd.DataFrame({
        "Ensembl_ID": df[id_col].iloc[idx_test].values,
        "True_Class": y_test,
        "Pred_Class": pred_class,
        "Prob_Pos":   pred_prob
    })
    save_csv(preds_df, os.path.join(out_dir, "test_predictions.csv"))

    model_info = pd.DataFrame({
        "Item":  ["Model","Positive_Class","CV_Method","CV_Folds","CV_Repeats","Best_mtry",
                  "Selected_Features_RFE","Selected_Features_Final"],
        "Value": ["RandomForest", pos_class, "repeatedcv","10","3",
                  getattr(gs.best_params_, "get", lambda k, d=None: None)("rf__max_features", None) if hasattr(gs, "best_params_") else None,
                  len(opt_vars), len(sel_vars)]
    })
    save_csv(model_info, os.path.join(out_dir, "model_summary.csv"))

    try:
        cv_summary = pd.DataFrame(gs.cv_results_)
        save_csv(cv_summary, os.path.join(out_dir, "cv_performance.csv"))
    except Exception:
        pass

    pd.DataFrame({"Feature": opt_vars}).to_csv(os.path.join(out_dir, "selected_features_rfe.csv"), index=False)
    pd.DataFrame({"Feature": sel_vars}).to_csv(os.path.join(out_dir, "selected_features_final.csv"), index=False)

    bundle = {
        "model": best_model,
        "preproc": preproc,
        "selected_vars": sel_vars,
        "all_vars_train": pred_cols,
        "id_col": id_col,
        "pos_class": pos_class,
        "class_levels": [pos_class, NEG_CLASS],
        "created_at": str(dt.datetime.now()),
        "session_info": {
            "python": sys.version,
            "platform": platform.platform(),
            "packages": {}  # می‌توانید با pip.freeze پر کنید
        }
    }
    joblib.dump(bundle, os.path.join(out_dir, "model_bundle.pkl"))
    _message("✅ ذخیره شد:", out_dir)
    return {"out_dir": out_dir, "best_model": best_model, "selected_vars": sel_vars, "preproc": preproc}

def step_ml_over_balanced_csvs(viruses=VIRUSES, out_subdir=OUT_SUBDIR, bal_subdir=BAL_SUBDIR):
    for virus in viruses:
        bal_dir = os.path.join(ROOT_RAW, virus, out_subdir, bal_subdir)
        if not os.path.isdir(bal_dir):
            _message("⏭️ مسیر نیست:", bal_dir); continue
        bal_files = [p for p in glob.glob(os.path.join(bal_dir, "*_balanced_top_vs_bottom.csv"))] + \
                    [p for p in glob.glob(os.path.join(bal_dir, f"*__top{TARGET_N}_HDF.csv")) if re.search(r"_bottom\d+_HRF__top\d+_HDF\.csv$", p, flags=re.I)]
        if not bal_files:
            _message("⏭️ فایل بالانس‌شده‌ای در", bal_dir, "پیدا نشد."); continue
        _message(f"====== ML → {virus} | files: {len(bal_files)} ======")
        for i, bal_file in enumerate(bal_files, 1):
            _message(f"({i}/{len(bal_files)}) پردازش:", os.path.basename(bal_file))
            try:
                df_bal = pd.read_csv(bal_file)
            except Exception as e:
                _message("❌ خطا در خواندن:", e); gc.collect(); continue
            if "Class" not in df_bal.columns:
                _message("⚠️ ستون Class یافت نشد:", os.path.basename(bal_file)); gc.collect(); continue
            if df_bal["Class"].nunique() < 2:
                _message("⚠️ کلاس کافی نیست. رد شد:", os.path.basename(bal_file)); gc.collect(); continue
            try:
                run_ml_and_export_csv(df=df_bal, balanced_csv_path=bal_file, pos_class=POS_CLASS)
                gc.collect()
            except Exception as e:
                _message("❌ خطای مدل‌سازی/خروجی:", e); gc.collect()
    _message("🎉 همهٔ فایل‌های بالانس‌شده پردازش شدند.")

# ============================================================

def _read_if_exists_csv(path, **kwargs):
    if os.path.isfile(path):
        try:
            return pd.read_csv(path, **kwargs)
        except Exception:
            return None
    return None

def step_merge_metrics():
    SUMMARY_DIR = os.path.join(ROOT_PUB, "_ML_SUMMARY")
    os.makedirs(SUMMARY_DIR, exist_ok=True)
    rows_wide = []
    rows_long = []

    for virus in VIRUSES:
        base_dir = os.path.join(ROOT_PUB, virus, OUT_SUBDIR, BAL_SUBDIR, "ML_results")
        if not os.path.isdir(base_dir):
            _message("⏭️ مسیر نیست:", base_dir); continue
        exp_dirs = [d for d in glob.glob(os.path.join(base_dir, "*")) if os.path.isdir(d)]
        if not exp_dirs:
            _message("⏭️ آزمایشی در", base_dir, "یافت نشد."); continue
        _message(f"====== Summary → {virus} | experiments: {len(exp_dirs)} ======")
        for exp_dir in exp_dirs:
            base_stub = os.path.basename(exp_dir)
            exp_name  = pretty_name(base_stub)
            f_metrics = os.path.join(exp_dir, "metrics_test.csv")
            met = _read_if_exists_csv(f_metrics)
            if met is None or not set(["Metric","Value"]).issubset(met.columns):
                _message("⚠️ metrics_test.csv موجود نیست/ناقص:", exp_dir); continue
            long_dt = met.copy()
            long_dt["Value"] = pd.to_numeric(long_dt["Value"], errors="coerce")
            long_dt.insert(0, "Virus", virus)
            long_dt.insert(1, "Experiment", exp_name)
            long_dt["Exp_Path"] = exp_dir
            rows_long.append(long_dt[["Virus","Experiment","Metric","Value","Exp_Path"]])

            vals = {row["Metric"]: row["Value"] for _, row in long_dt.iterrows()}
            wide_dt = pd.DataFrame([vals])
            wide_dt.insert(0, "Exp_Path", exp_dir)
            wide_dt.insert(0, "Experiment", exp_name)
            wide_dt.insert(0, "Virus", virus)
            rows_wide.append(wide_dt)

    if rows_long:
        ALL_LONG = pd.concat(rows_long, ignore_index=True)
        ALL_LONG = ALL_LONG.sort_values(["Virus","Experiment","Metric"])
        out_long = os.path.join(SUMMARY_DIR, "all_models_metrics_long.csv")
        save_csv(ALL_LONG, out_long)
        _message("✅ ذخیره شد:", out_long)
    else:
        _message("⚠️ هیچ metrics_test.csv برای ساخت long پیدا نشد.")

    if rows_wide:
        ALL_WIDE = pd.concat(rows_wide, ignore_index=True).sort_values(["Virus","Experiment"])
        out_wide = os.path.join(SUMMARY_DIR, "all_models_metrics_wide.csv")
        save_csv(ALL_WIDE, out_wide)
        _message("✅ ذخیره شد:", out_wide)
    else:
        _message("⚠️ هیچ metrics_test.csv برای ساخت wide پیدا نشد.")

# ============================================================
# 5) Prediction روی کل ژنوم + Excel (TopN) + UpSet
# ============================================================
def step_predict_full_genome():
    PRED_DIR   = os.path.join(ROOT_PUB, "_PREDICTIONS")
    os.makedirs(PRED_DIR, exist_ok=True)
    XLSX_PATH  = os.path.join(PRED_DIR, "all_models_predictions.xlsx")
    UPSET_PATH = os.path.join(PRED_DIR, "upset_top1000.png")

    WRITE_MODE = "TOP_ONLY"   # یا "FULL_TO_EXCEL"
    TOP_N_XLSX = 1000
    BATCH_SIZE = 5000

    def read_if(p):
        if os.path.isfile(p):
            try:
                return pd.read_csv(p)
            except Exception:
                return None
        return None

    feat_all = read_if(FEATURES_CSV)
    if feat_all is None:
        raise RuntimeError("عدم توانایی در خواندن FEATURES_CSV")
    id_candidates = ["Ensembl_ID","ENSEMBL_ID","EnsemblId","ensembl_id"]
    id_col = next((c for c in id_candidates if c in feat_all.columns), None)
    if id_col is None:
        raise RuntimeError("ستون شناسه Ensembl_ID یافت نشد در فیچرها.")
    feat_all = feat_all.drop_duplicates(subset=[id_col]).reset_index(drop=True)

    num_cols_all = [c for c in feat_all.columns if c != id_col]
    feat_num = feat_all.copy()
    for c in num_cols_all:
        feat_num[c] = pd.to_numeric(feat_num[c], errors="coerce")

    bundle_files = []
    for virus in VIRUSES:
        base = os.path.join(ROOT_PUB, virus, OUT_SUBDIR, BAL_SUBDIR, "ML_results")
        if not os.path.isdir(base):
            continue
        subdirs = [d for d in glob.glob(os.path.join(base, "*")) if os.path.isdir(d)]
        for sd in subdirs:
            bf = os.path.join(sd, "model_bundle.pkl")
            if os.path.isfile(bf):
                bundle_files.append(bf)
    if not bundle_files:
        raise RuntimeError("هیچ model_bundle.pkl پیدا نشد.")

    wb = openpyxl.Workbook()
    if "Sheet" in [ws.title for ws in wb.worksheets]:
        wb.remove(wb["Sheet"])
    ws_summary = wb.create_sheet("SUMMARY")
    summary_rows = []
    top_sets = {}

    for bf in bundle_files:
        bundle = joblib.load(bf)
        if not all(k in bundle for k in ["model","preproc","pos_class"]):
            raise RuntimeError("باندل ناقص است: " + bf)

        exp_dir   = os.path.dirname(bf)
        exp_name0 = os.path.basename(exp_dir)
        exp_name  = pretty_name(exp_name0)
        sheet     = make_unique_sheet(wb, exp_name)

        if bundle.get("all_vars_train"):
            pre_cols = bundle["all_vars_train"]
        else:
            pre_cols = num_cols_all

        missing_cols = [c for c in pre_cols if c not in feat_num.columns]
        if missing_cols:
            raise RuntimeError(f"این فیچرها در combinedFinal.csv نیستند ({len(missing_cols)}): {', '.join(missing_cols)}")

        n = feat_num.shape[0]
        posc = bundle["pos_class"]
        pred_prob = np.zeros(n, dtype=float)

        sel_vars = bundle.get("selected_vars") or pre_cols

        for i in range(0, n, BATCH_SIZE):
            idx = slice(i, min(i + BATCH_SIZE, n))
            raw_full = feat_num.loc[idx, pre_cols].copy()
            X_full = bundle["preproc"].transform(raw_full)
            sel_idx = [pre_cols.index(v) for v in sel_vars]
            predX = X_full[:, sel_idx]
            model = bundle["model"]
            try:
                # imblearn pipeline?
                classes_model = getattr(model[-1], "classes_", None)
            except Exception:
                classes_model = None
            if classes_model is None:
                try:
                    classes_model = getattr(model.named_steps["rf"], "classes_", None)
                except Exception:
                    classes_model = np.array([NEG_CLASS, POS_CLASS])
            pos_index = list(classes_model).index(posc) if posc in classes_model else 1
            prob_batch = model.predict_proba(predX)[:, pos_index]
            pred_prob[idx] = prob_batch
            del raw_full, X_full, predX; gc.collect()

        res = pd.DataFrame({
            "Ensembl_ID": feat_num[id_col].values,
            "Prob_HDF": pred_prob
        })
        res = res.sort_values(["Prob_HDF","Ensembl_ID"], ascending=[False, True]).reset_index(drop=True)
        res["Rank"] = np.arange(1, res.shape[0]+1)

        tsv_path = os.path.join(PRED_DIR, f"{sanitize_sheet(exp_name)}_predictions.tsv.gz")
        with gzip.open(tsv_path, "wt") as gz:
            res.to_csv(gz, sep="\t", index=False)

        ws = wb.create_sheet(sheet)
        if WRITE_MODE == "FULL_TO_EXCEL":
            head = res
        else:
            head = res.head(TOP_N_XLSX)
        for j, col in enumerate(head.columns, 1):
            ws.cell(row=1, column=j, value=col)
        for i, row in enumerate(head.itertuples(index=False), 2):
            for j, v in enumerate(row, 1):
                ws.cell(row=i, column=j, value=float(v) if isinstance(v, (np.floating, float)) else (int(v) if isinstance(v, (np.integer, int)) else v))

        summary_rows.append({
            "Experiment": exp_name,
            "Sheet": sheet,
            "ModelRDS": bf,
            "TSV_GZ": tsv_path,
            "N_all": res.shape[0],
            "TopN_in_Excel": res.shape[0] if WRITE_MODE=="FULL_TO_EXCEL" else min(TOP_N_XLSX, res.shape[0]),
            "Top1000_MinProb": float(res["Prob_HDF"].iloc[999]) if res.shape[0] >= 1000 else np.nan
        })

        top_sets[sheet] = set(res["Ensembl_ID"].head(1000).astype(str).tolist())
        del res, pred_prob; gc.collect()

    cols = ["Experiment","Sheet","ModelRDS","TSV_GZ","N_all","TopN_in_Excel","Top1000_MinProb"]
    for j, col in enumerate(cols, 1):
        ws_summary.cell(row=1, column=j, value=col)
    for i, row in enumerate(summary_rows, 2):
        for j, col in enumerate(cols, 1):
            ws_summary.cell(row=i, column=j, value=row[col])

    wb.save(XLSX_PATH)
    _message("✅ Excel saved:", XLSX_PATH)

    if len(top_sets) >= 2 and _HAS_UPSET:
        contents = {k: v for k, v in top_sets.items()}
        inc = from_contents(contents)
        plt.figure(figsize=(12,7))
        UpSet(inc, subset_size='count', sort_by='cardinality', intersection_plot_elements=10).plot()
        plt.tight_layout()
        plt.savefig(UPSET_PATH, dpi=150)
        plt.close()
        _message("✅ UpSet saved:", UPSET_PATH)
    else:
        _message("⚠️ UpSet ساخته نشد (کمتر از ۲ مدل یا نبود upsetplot).")

    _message("✅ پیش‌بینی کل ژنوم تمام شد.")

# ============================================================
# Main: اجرای سرراست همه مراحل
# ============================================================
def run_all():
    step_build_top_bottom()
    step_label_and_balance()
    step_ml_over_balanced_csvs()
    # step_merge_metrics()
    # step_predict_full_genome()

if __name__ == "__main__":
    run_all()

